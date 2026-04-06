"""
extractor_igss.py
=================
Fase 1 del pipeline ETL — Extracción de datos del Excel oficial del IGSS.

Extrae las hojas clave del archivo IGSS_en_Cifras_2025.xlsm y las guarda
como CSV en data/processed/ para su posterior limpieza y carga al DW.

Uso:
    python src/ingesta/extractor_igss.py

Requisitos:
    pip install openpyxl pandas loguru
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from loguru import logger

# ── Configuración ───────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parents[2]
RAW_DIR = BASE_DIR / "data" / "raw"
PROCESSED_DIR = BASE_DIR / "data" / "processed"
LOG_DIR = BASE_DIR / "data" / "logs"

EXCEL_FILE = RAW_DIR / "IGSS_en_Cifras_2025.xlsm"

# Crear directorios si no existen
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)

# Configurar logging
logger.add(
    LOG_DIR / "extraccion_{time}.log",
    rotation="1 MB",
    retention="30 days",
    level="INFO"
)

# ── Mapa de departamentos ────────────────────────────────────────────────────
DEPTO_ID_MAP = {
    "Guatemala": 1, "El Progreso": 2, "Sacatepéquez": 3,
    "Chimaltenango": 4, "Escuintla": 5, "Santa Rosa": 6,
    "Sololá": 7, "Totonicapán": 8, "Quetzaltenango": 9,
    "Suchitepéquez": 10, "Retalhuleu": 11, "San Marcos": 12,
    "Huehuetenango": 13, "Quiché": 14, "Baja Verapaz": 15,
    "Alta Verapaz": 16, "Petén": 17, "Izabal": 18,
    "Zacapa": 19, "Chiquimula": 20, "Jalapa": 21,
    "Jutiapa": 22, "Multiregional": 30
}

DEPTO_REGION_MAP = {
    1: "Metropolitana", 2: "Nororiente", 3: "Central", 4: "Central",
    5: "Sur", 6: "Suroriente", 7: "Suroccidente", 8: "Suroccidente",
    9: "Suroccidente", 10: "Suroccidente", 11: "Suroccidente",
    12: "Suroccidente", 13: "Noroccidente", 14: "Noroccidente",
    15: "Norte", 16: "Norte", 17: "Petén", 18: "Nororiente",
    19: "Nororiente", 20: "Nororiente", 21: "Suroriente", 22: "Suroriente"
}


def _is_formula(val) -> bool:
    """Verifica si un valor es una fórmula Excel sin calcular."""
    return isinstance(val, str) and val.strip().startswith("=")


def _safe_float(val) -> float | None:
    """Convierte un valor a float de forma segura."""
    if val is None or _is_formula(val):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


class IGSSExtractor:
    """Extractor principal para el Excel del IGSS."""

    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self.wb = None
        self.ingestion_log = {
            "timestamp": datetime.now().isoformat(),
            "archivo_fuente": str(excel_path),
            "extracciones": []
        }

    def cargar_workbook(self):
        """Carga el workbook de Excel (modo solo lectura)."""
        logger.info(f"Cargando workbook: {self.excel_path}")
        if not self.excel_path.exists():
            raise FileNotFoundError(
                f"No se encontró el archivo: {self.excel_path}\n"
                "Descárgalo desde https://www.igssgt.org y colócalo en data/raw/"
            )
        self.wb = load_workbook(
            self.excel_path,
            read_only=True,
            keep_vba=True,
            data_only=True  # Obtiene valores calculados, no fórmulas
        )
        logger.success(f"Workbook cargado. Hojas disponibles: {len(self.wb.sheetnames)}")

    def extraer_costos_unitarios(self) -> pd.DataFrame:
        """
        Extrae la tabla H1: Costo medio unitario por tipo de servicio.
        Período 2014-2024.
        """
        logger.info("Extrayendo Hoja H1: Costos Unitarios...")
        ws = self.wb["H1"]

        registros = []
        encabezado_encontrado = False

        for row in ws.iter_rows(max_row=25, values_only=True):
            vals = [v for v in row if v is not None]
            if not vals:
                continue

            # Detectar fila de encabezado
            if vals[0] == "Año":
                encabezado_encontrado = True
                continue

            if encabezado_encontrado and isinstance(vals[0], int) and 2000 <= vals[0] <= 2030:
                anio = vals[0]
                hosp = _safe_float(vals[1]) if len(vals) > 1 else None
                ce = _safe_float(vals[2]) if len(vals) > 2 else None
                emerg = _safe_float(vals[3]) if len(vals) > 3 else None
                pa = _safe_float(vals[4]) if len(vals) > 4 else None

                # Excluir años con datos incompletos (fórmulas sin calcular)
                if hosp is not None and ce is not None:
                    registros.append({
                        "anio": anio,
                        "hospitalizacion_q": round(hosp, 2),
                        "consulta_externa_q": round(ce, 2),
                        "emergencia_q": round(emerg, 2) if emerg else None,
                        "primeros_auxilios_q": round(pa, 2) if pa else None
                    })

        df = pd.DataFrame(registros)
        logger.success(f"  → {len(df)} registros extraídos de H1")
        self._log_extraccion("H1_costos_unitarios", len(df), "costos_unitarios.csv")
        return df

    def extraer_ejecucion_gastos(self) -> pd.DataFrame:
        """
        Extrae la tabla Ejec. y Gastos: Egresos e Ingresos por departamento.
        Años 2024 y 2025.
        """
        logger.info("Extrayendo Hoja 'Ejec. y Gastos'...")
        ws = self.wb["Ejec. y Gastos"]

        registros = []
        anios = [2024, 2025]

        for row in ws.iter_rows(max_row=50, values_only=True):
            vals = [v for v in row]
            if len(vals) < 4:
                continue

            # Detectar filas de datos: columna 1 = ID numérico
            if isinstance(vals[0], (int, float)) and vals[0] in range(1, 31):
                id_depto = int(vals[0])
                nombre = vals[1]
                egresos_2024 = _safe_float(vals[2])
                ingresos_2024 = _safe_float(vals[3])

                if nombre and egresos_2024 is not None:
                    registros.append({
                        "id_departamento": id_depto,
                        "departamento": nombre,
                        "anio": 2024,
                        "egresos_q": round(egresos_2024, 2),
                        "ingresos_q": round(ingresos_2024, 2) if ingresos_2024 else 0.0
                    })

                # Columna 5-8: datos 2025
                if len(vals) >= 8 and isinstance(vals[4], (int, float)):
                    egresos_2025 = _safe_float(vals[6])
                    ingresos_2025 = _safe_float(vals[7])

                    if egresos_2025 is not None:
                        registros.append({
                            "id_departamento": id_depto,
                            "departamento": nombre,
                            "anio": 2025,
                            "egresos_q": round(egresos_2025, 2),
                            "ingresos_q": round(ingresos_2025, 2) if ingresos_2025 else 0.0
                        })

        df = pd.DataFrame(registros)

        # Calcular métricas derivadas
        df["brecha_q"] = df["egresos_q"] - df["ingresos_q"]
        df["ratio_ei"] = (df["egresos_q"] / df["ingresos_q"].replace(0, float("nan"))).round(4)
        df["region"] = df["id_departamento"].map(DEPTO_REGION_MAP)
        df["flag_anomalia"] = df["ratio_ei"] > 4.0

        logger.success(f"  → {len(df)} registros extraídos de Ejec. y Gastos")
        self._log_extraccion("Ejec_y_Gastos", len(df), "ejecucion_gastos.csv")
        return df

    def extraer_costos_departamento(self) -> pd.DataFrame:
        """
        Extrae la tabla 5.1 TD_Costos: Costos mensuales por departamento.
        Años 2021-2025.
        """
        logger.info("Extrayendo Hoja '5.1 TD_Costos'...")
        ws = self.wb["5.1 TD_Costos"]

        meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        anios_encontrados = []
        registros = []
        mes_cols = {}  # {col_index: (anio, mes)}

        # Primera pasada: detectar estructura de columnas
        fila_anio = None
        fila_mes = None
        fila_datos_inicio = None

        for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
            vals = list(row)
            # Buscar fila de años
            anios_en_fila = [v for v in vals if isinstance(v, int) and 2020 <= v <= 2030]
            if len(anios_en_fila) >= 3:
                fila_anio = i
                anios_encontrados = [(j, v) for j, v in enumerate(vals)
                                     if isinstance(v, int) and 2020 <= v <= 2030]

            # Buscar fila con "Id."
            if "Id." in str(vals):
                fila_datos_inicio = i + 1
                break

        registros = []
        datos_iniciados = False

        for i, row in enumerate(ws.iter_rows(max_row=200, values_only=True)):
            vals = list(row)

            if fila_datos_inicio and i < fila_datos_inicio:
                continue

            if not vals or vals[0] is None:
                continue

            if isinstance(vals[0], (int, float)) and 1 <= int(vals[0]) <= 30:
                id_depto = int(vals[0])
                nombre = vals[1] if len(vals) > 1 else None
                if not nombre:
                    continue

                # Extraer costos por año/mes (columnas 2 en adelante)
                # Estructura: col2=Jan21, col3=Feb21, ... col13=Dec21, col14=Jan22, etc.
                col_offset = 2
                for anio in range(2021, 2026):
                    for j, mes in enumerate(meses):
                        col_idx = col_offset + (anio - 2021) * 12 + j
                        if col_idx < len(vals):
                            costo = _safe_float(vals[col_idx])
                            if costo is not None and costo > 0:
                                registros.append({
                                    "id_departamento": id_depto,
                                    "departamento": nombre,
                                    "anio": anio,
                                    "mes": j + 1,
                                    "nombre_mes": mes,
                                    "costo_total_q": round(costo, 2),
                                    "region": DEPTO_REGION_MAP.get(id_depto, "Sin clasificar")
                                })

        df = pd.DataFrame(registros)
        logger.success(f"  → {len(df)} registros extraídos de 5.1 TD_Costos")
        self._log_extraccion("5.1_TD_Costos", len(df), "costos_departamento.csv")
        return df

    def _log_extraccion(self, hoja: str, registros: int, archivo_destino: str):
        """Registra metadata de cada extracción."""
        self.ingestion_log["extracciones"].append({
            "hoja_origen": hoja,
            "registros_extraidos": registros,
            "archivo_destino": archivo_destino,
            "timestamp": datetime.now().isoformat()
        })

    def guardar_log(self):
        """Guarda el log de ingesta en JSON."""
        log_path = LOG_DIR / f"ingestion_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump(self.ingestion_log, f, indent=2, ensure_ascii=False)
        logger.info(f"Log de ingesta guardado: {log_path}")


def main():
    logger.info("=" * 60)
    logger.info("IGSS BI — Fase 1: Extracción de Datos")
    logger.info("=" * 60)

    extractor = IGSSExtractor(EXCEL_FILE)

    try:
        extractor.cargar_workbook()

        # Extracción 1: Costos Unitarios
        df_costos = extractor.extraer_costos_unitarios()
        out_path = PROCESSED_DIR / "costos_unitarios.csv"
        df_costos.to_csv(out_path, index=False, encoding="utf-8")
        logger.success(f"Guardado: {out_path}")

        # Extracción 2: Ejecución y Gastos
        df_ejec = extractor.extraer_ejecucion_gastos()
        out_path = PROCESSED_DIR / "ejecucion_gastos.csv"
        df_ejec.to_csv(out_path, index=False, encoding="utf-8")
        logger.success(f"Guardado: {out_path}")

        # Extracción 3: Costos por Departamento
        df_deptos = extractor.extraer_costos_departamento()
        out_path = PROCESSED_DIR / "costos_departamento.csv"
        df_deptos.to_csv(out_path, index=False, encoding="utf-8")
        logger.success(f"Guardado: {out_path}")

        # Guardar log
        extractor.guardar_log()

        logger.info("=" * 60)
        logger.success("✅ Extracción completada exitosamente")
        logger.info(f"  Costos unitarios: {len(df_costos)} registros")
        logger.info(f"  Ejecución gastos: {len(df_ejec)} registros")
        logger.info(f"  Costos x depto:  {len(df_deptos)} registros")
        logger.info("=" * 60)

    except FileNotFoundError as e:
        logger.error(str(e))
        sys.exit(1)
    except Exception as e:
        logger.exception(f"Error inesperado: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
